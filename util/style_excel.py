from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
import openpyxl
from openpyxl.comments import Comment
from copy import copy


class ExcelStyle:
    def format_excel(self, route, comment):
        print("Ingresa a formatear estilos de Excel")
        wb = load_workbook(route)
        ws = wb.active
        self.ws = ws

        self.add_comment_cell(ws, comment)

        self.insert_image(ws)

        self.format_title(ws)

        self.column_spacing(ws)

        self.center_data(ws)

        self.format_HH_MM_SS(ws,7,19)

        # Guardado del nuevo Archivo Formateado
        print("Finaliza formateo de Archivo Excel Ruta: ",route)
        wb.save(route)

    def format_title(self, ws):
        for row in ws.iter_rows(min_row=6, max_col=ws.max_column, max_row=6):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FF6464")
                cell.font = Font(b=True, color="FFFFE0")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.auto_filter.ref = "A6:Z6"


    def insert_image(self, ws):
        # Se agrega 5 filas
        ws.insert_rows(1, 5)

        ws.merge_cells('A1:Z5')
        img = openpyxl.drawing.image.Image('assets/mct.png')
        img.anchor = 'A1'
        ws.add_image(img)

    def column_spacing(self, ws):
        dims = {}
        for row in ws.columns:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 1), len(str(cell.value))))

        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 2


    def center_data(self, ws):
        for col in ws.columns:
            for cell in col:
                alignment_obj = copy(cell.alignment)
                alignment_obj.horizontal = 'center'
                alignment_obj.vertical = 'center'
                cell.alignment = alignment_obj

    def format_HH_MM_SS(self, ws, start_column, start_row):
        nsmmyy = NamedStyle(name="cd1", number_format="HH:MM:SS")
        for row in ws[start_column:ws.max_row]:
            for cell in row[start_row:ws.max_column]:
                cell.style = nsmmyy

    def add_comment_cell(self, ws, comment):
        for i in range(ws.max_row):
            if i > 1:
                comment_cell = ws.cell(row = i, column = 2).value
                message_wrote_all = comment[comment_cell]

                message = message_wrote_all[('message_wrote', 'LLEGADA A PLANTA')]
                if (type(message) == str) & (message != "") & (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row = i, column = 9).comment = comments

                message = message_wrote_all[('message_wrote', 'INGRESO A PLANTA')]
                if (type(message) == str) & (message != "")& (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row=i, column=10).comment = comments

                message = message_wrote_all[('message_wrote', 'INICIO DEL CARGUE')]
                if (type(message) == str) & (message != "")& (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row=i, column=11).comment = comments

                message = message_wrote_all[('message_wrote', 'FIN DEL CARGUE')]
                if (type(message) == str) & (message != "")& (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row=i, column=12).comment = comments


                message = message_wrote_all[('message_wrote', 'SALIDA DE PLANTA')]
                if (type(message) == str) & (message != "")& (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row=i, column=13).comment = comments

                message = message_wrote_all[('message_wrote', 'LLEGADA A DESCARGUE')]
                if (type(message) == str) & (message != "")& (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row=i, column=14).comment = comments

                message = message_wrote_all[('message_wrote', 'INGRESO A DESCARGUE')]
                if (type(message) == str) & (message != "")& (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row=i, column=15).comment = comments

                message = message_wrote_all[('message_wrote', 'INICIO DEL DESCARGUE')]
                if (type(message) == str) & (message != "")& (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row=i, column=16).comment = comments

                message = message_wrote_all[('message_wrote', 'FIN DEL DESCARGUE')]
                if (type(message) == str) & (message != "")& (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row=i, column=17).comment = comments

                message = message_wrote_all[('message_wrote', 'SALIDA DE DESCARGUE')]
                if (type(message) == str) & (message != "")& (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row=i, column=18).comment = comments

                message = message_wrote_all[('message_wrote', 'VEHICULO RETORNADO A PLANTA POLAR')]
                if (type(message) == str) & (message != "")& (message != ".") & (message != ".."):
                    comments = Comment(message, "Funcionario MCT")
                    ws.cell(row=i, column=19).comment = comments



